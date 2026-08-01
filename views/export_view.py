"""
Export & Version Management Views
Supports: Excel, PNG, PDF, CSV (FMEA), version snapshots.
"""
import os
import json
from datetime import datetime

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QGroupBox, QMessageBox, QFileDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QTextEdit, QLineEdit, QFormLayout,
    QComboBox, QFrame
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QColor, QFont

from engines.compute import HOQEngine
from views.styles import COLORS, PHASE_NAMES, DIRECTION_SYMBOLS
from utils.i18n import t


class ExportView(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.project_id = None
        self.phase = 1
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        header = QHBoxLayout()
        lbl = QLabel(t("export.title"))
        lbl.setObjectName("title")
        header.addWidget(lbl)
        layout.addLayout(header)

        # Export options
        excel_group = QGroupBox(t("export.excel_group"))
        excel_layout = QVBoxLayout(excel_group)
        excel_desc = QLabel(t("export.excel_desc"))
        excel_desc.setWordWrap(True)
        excel_layout.addWidget(excel_desc)
        btn_excel = QPushButton(t("export.excel"))
        btn_excel.clicked.connect(self._export_excel)
        btn_excel.setMinimumHeight(40)
        excel_layout.addWidget(btn_excel)
        layout.addWidget(excel_group)

        png_group = QGroupBox(t("export.png_group"))
        png_layout = QVBoxLayout(png_group)
        png_desc = QLabel(t("export.png_desc"))
        png_desc.setWordWrap(True)
        png_layout.addWidget(png_desc)
        btn_png = QPushButton(t("export.png"))
        btn_png.clicked.connect(self._export_png)
        btn_png.setMinimumHeight(40)
        png_layout.addWidget(btn_png)
        layout.addWidget(png_group)

        fmea_group = QGroupBox(t("export.fmea_group"))
        fmea_layout = QVBoxLayout(fmea_group)
        fmea_desc = QLabel(t("export.fmea_desc"))
        fmea_desc.setWordWrap(True)
        fmea_layout.addWidget(fmea_desc)
        btn_fmea = QPushButton(t("export.fmea"))
        btn_fmea.clicked.connect(self._export_fmea_csv)
        btn_fmea.setMinimumHeight(40)
        fmea_layout.addWidget(btn_fmea)
        layout.addWidget(fmea_group)

        json_group = QGroupBox(t("export.json_group"))
        json_layout = QVBoxLayout(json_group)
        json_desc = QLabel(t("export.json_desc"))
        json_desc.setWordWrap(True)
        json_layout.addWidget(json_desc)
        btn_json = QPushButton(t("export.json"))
        btn_json.clicked.connect(self._export_json)
        btn_json.setMinimumHeight(40)
        json_layout.addWidget(btn_json)
        layout.addWidget(json_group)

        layout.addStretch()

    def set_project(self, project_id, phase=1):
        self.project_id = project_id
        self.phase = phase

    def _export_excel(self):
        if not self.project_id:
            return
        path, _ = QFileDialog.getSaveFileName(self, t("export.dlg_excel_title"), "QFD_Report.xlsx",
                                              t("export.excel_filter"))
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()

            # Sheet 1: VOC
            ws_voc = wb.active
            ws_voc.title = t("export.sheet_voc")
            vocs = self.db.get_vocs(self.project_id, self.phase)
            headers_v = [t("export.h_id"), t("export.h_name"), t("export.h_desc"), t("export.h_source"),
                         t("export.h_importance"), t("export.h_kano_type"), t("export.h_improvement"), t("export.h_sales_point")]
            for j, h in enumerate(headers_v, 1):
                cell = ws_voc.cell(row=1, column=j, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1B5E8C")
                cell.alignment = Alignment(horizontal='center')
            for i, v in enumerate(vocs, 2):
                ws_voc.cell(row=i, column=1, value=v['id'])
                ws_voc.cell(row=i, column=2, value=v['name'])
                ws_voc.cell(row=i, column=3, value=v['description'] or '')
                ws_voc.cell(row=i, column=4, value=v['source'] or '')
                ws_voc.cell(row=i, column=5, value=v['importance'])
                ws_voc.cell(row=i, column=6, value=v['kano_type'] or '')
                ws_voc.cell(row=i, column=7, value=v['improvement_ratio'])
                ws_voc.cell(row=i, column=8, value=v['sales_point'])

            # Sheet 2: CTQ
            ws_ctq = wb.create_sheet(t("export.sheet_ctq"))
            ctqs = self.db.get_ctqs(self.project_id, self.phase)
            headers_c = [t("export.h_id"), t("export.h_ctq_name"), t("export.h_unit"), t("export.h_direction"),
                         t("export.h_target"), t("export.h_current"), t("export.h_difficulty")]
            for j, h in enumerate(headers_c, 1):
                cell = ws_ctq.cell(row=1, column=j, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1B5E8C")
            for i, c in enumerate(ctqs, 2):
                ws_ctq.cell(row=i, column=1, value=c['id'])
                ws_ctq.cell(row=i, column=2, value=c['name'])
                ws_ctq.cell(row=i, column=3, value=c['unit'] or '')
                ws_ctq.cell(row=i, column=4, value=c['direction'])
                ws_ctq.cell(row=i, column=5, value=c['target_value'] or '')
                ws_ctq.cell(row=i, column=6, value=c['current_value'] or '')
                ws_ctq.cell(row=i, column=7, value=c['difficulty'])

            # Sheet 3: Relationship matrix
            ws_rel = wb.create_sheet(t("export.sheet_rel"))
            # Headers
            ws_rel.cell(row=1, column=1, value=t("export.h_voc_vs_ctq")).font = Font(bold=True)
            for j, c in enumerate(ctqs, 2):
                ws_rel.cell(row=1, column=j, value=c['name']).font = Font(bold=True)
            for i, v in enumerate(vocs, 2):
                ws_rel.cell(row=i, column=1, value=v['name']).font = Font(bold=True)
                for j, c in enumerate(ctqs, 2):
                    rel = self.db.get_relationship(self.project_id, v['id'], c['id'], self.phase)
                    val = rel['strength'] if rel else 0
                    if val:
                        cell = ws_rel.cell(row=i, column=j, value=val)
                        if val == 9:
                            cell.fill = PatternFill("solid", fgColor="D4E6F1")
                        elif val == 3:
                            cell.fill = PatternFill("solid", fgColor="EBF5FB")

            # Sheet 4: Importance results
            ws_imp = wb.create_sheet(t("export.sheet_importance"))
            rels_data = [dict(r) for r in self.db.get_relationships(self.project_id, self.phase)]
            importance = HOQEngine.compute_importance(
                [dict(v) for v in vocs], [dict(c) for c in ctqs], rels_data)
            headers_i = [t("export.h_ctq_name"), t("export.h_abs_importance"), t("export.h_rel_importance"),
                         t("export.h_weighted_importance"), t("export.h_rank")]
            for j, h in enumerate(headers_i, 1):
                cell = ws_imp.cell(row=1, column=j, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="E67E22")
            for i, c in enumerate(ctqs, 2):
                imp = importance.get(c['id'], {})
                ws_imp.cell(row=i, column=1, value=c['name'])
                ws_imp.cell(row=i, column=2, value=round(imp.get('absolute_importance', 0), 2))
                ws_imp.cell(row=i, column=3, value=round(imp.get('relative_importance', 0), 1))
                ws_imp.cell(row=i, column=4, value=round(imp.get('weighted_importance', 0), 2))
                ws_imp.cell(row=i, column=5, value=imp.get('rank', 0))

            # Auto-fit columns
            for ws in [ws_voc, ws_ctq, ws_rel, ws_imp]:
                for col in ws.columns:
                    max_length = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

            wb.save(path)
            QMessageBox.information(self, t("common.success"), t("export.excel_success", path=path))
        except Exception as e:
            QMessageBox.critical(self, t("export.export_failed"), str(e))

    def _export_png(self):
        if not self.project_id:
            return
        path, _ = QFileDialog.getSaveFileName(self, t("export.dlg_png_title"), "QFD_Matrix.png",
                                              t("export.png_filter"))
        if not path:
            return
        try:
            from utils.fonts import setup_matplotlib_fonts
            setup_matplotlib_fonts()
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_agg import FigureCanvasAgg

            vocs = [dict(v) for v in self.db.get_vocs(self.project_id, self.phase)]
            ctqs = [dict(c) for c in self.db.get_ctqs(self.project_id, self.phase)]
            rels = [dict(r) for r in self.db.get_relationships(self.project_id, self.phase)]

            nv, nc = len(vocs), len(ctqs)
            if nv == 0 or nc == 0:
                QMessageBox.warning(self, t("common.hint"), t("export.empty_matrix"))
                return

            # Create matrix visualization
            fig = Figure(figsize=(max(10, nc * 0.8 + 4), max(6, nv * 0.5 + 3)), dpi=150)
            canvas = FigureCanvasAgg(fig)
            ax = fig.add_subplot(111)

            # Build data
            rel_map = {(r['voc_id'], r['ctq_id']): r['strength'] for r in rels}
            data = []
            for v in vocs:
                row = []
                for c in ctqs:
                    row.append(rel_map.get((v['id'], c['id']), 0))
                data.append(row)

            import numpy as np
            data_arr = np.array(data)
            im = ax.imshow(data_arr, cmap='Blues', aspect='auto', vmin=0, vmax=9)

            # Labels
            ax.set_xticks(range(nc))
            ax.set_xticklabels([c['name'] for c in ctqs], rotation=45, ha='right', fontsize=8)
            ax.set_yticks(range(nv))
            ax.set_yticklabels([v['name'] for v in vocs], fontsize=8)

            # Cell text
            symbol_map = {9: "●", 3: "◎", 1: "△", 0: ""}
            for i in range(nv):
                for j in range(nc):
                    val = data_arr[i, j]
                    text = symbol_map.get(int(val), str(int(val)) if val else "")
                    color = 'white' if val >= 5 else 'black'
                    ax.text(j, i, text, ha='center', va='center', fontsize=10, color=color)

            ax.set_title(t("export.png_title"), fontsize=14, fontweight='bold', pad=15)
            fig.colorbar(im, ax=ax, shrink=0.6, label=t("export.png_colorbar"))
            fig.tight_layout()

            if path.endswith('.svg'):
                fig.savefig(path, format='svg', bbox_inches='tight')
            else:
                fig.savefig(path, dpi=200, bbox_inches='tight')

            QMessageBox.information(self, t("common.success"), t("export.png_success", path=path))
        except Exception as e:
            QMessageBox.critical(self, t("export.export_failed"), str(e))

    def _export_fmea_csv(self):
        if not self.project_id:
            return
        path, _ = QFileDialog.getSaveFileName(self, t("export.dlg_fmea_title"), "FMEA_CTQ.csv",
                                              t("common.csv_filter"))
        if not path:
            return
        try:
            import csv
            ctqs = [dict(c) for c in self.db.get_ctqs(self.project_id, self.phase)]
            vocs = [dict(v) for v in self.db.get_vocs(self.project_id, self.phase)]
            rels = [dict(r) for r in self.db.get_relationships(self.project_id, self.phase)]
            importance = HOQEngine.compute_importance(vocs, ctqs, rels)

            with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow([
                    t("export.h_ctq_id"), t("export.h_ctq_name"), t("export.h_unit"), t("export.h_target"), t("export.h_current"),
                    t("export.h_direction"), t("export.h_difficulty"), t("export.h_abs_importance"), t("export.h_rel_importance"),
                    t("export.h_rank"), t("export.h_is_key_ctq")
                ])
                n = len(ctqs)
                top20 = max(1, int(n * 0.2))
                for c in ctqs:
                    imp = importance.get(c['id'], {})
                    rank = imp.get('rank', 0)
                    writer.writerow([
                        c['id'], c['name'], c['unit'] or '',
                        c['target_value'] or '', c['current_value'] or '',
                        DIRECTION_SYMBOLS.get(c['direction'], ''),
                        c['difficulty'],
                        round(imp.get('absolute_importance', 0), 2),
                        round(imp.get('relative_importance', 0), 1),
                        rank,
                        t("export.yes") if rank <= top20 else t("export.no")
                    ])
            QMessageBox.information(self, t("common.success"), t("export.fmea_success", path=path))
        except Exception as e:
            QMessageBox.critical(self, t("export.export_failed"), str(e))

    def _export_json(self):
        if not self.project_id:
            return
        path, _ = QFileDialog.getSaveFileName(self, t("export.dlg_json_title"), "QFD_Project.json",
                                              t("msg.import_json_filter"))
        if not path:
            return
        try:
            data = self.db.export_project_data(self.project_id)
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            QMessageBox.information(self, t("common.success"), t("export.json_success", path=path))
        except Exception as e:
            QMessageBox.critical(self, t("export.export_failed"), str(e))


class VersionView(QWidget):
    version_restored = Signal()

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.project_id = None
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)

        header = QHBoxLayout()
        lbl = QLabel(t("ver.title"))
        lbl.setObjectName("title")
        header.addWidget(lbl)
        header.addStretch()

        btn_save = QPushButton(t("ver.save"))
        btn_save.setObjectName("success")
        btn_save.clicked.connect(self._save_version)
        header.addWidget(btn_save)
        layout.addLayout(header)

        hint = QLabel(t("ver.hint"))
        hint.setObjectName("subtitle")
        layout.addWidget(hint)

        # Label input
        label_layout = QHBoxLayout()
        label_layout.addWidget(QLabel(t("ver.label")))
        self.label_input = QLineEdit()
        self.label_input.setPlaceholderText(t("ver.label_placeholder"))
        label_layout.addWidget(self.label_input)
        layout.addLayout(label_layout)

        # Version list
        self.table = QTableWidget()
        self.table.verticalHeader().setDefaultSectionSize(32)
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels([t("ver.col_id"), t("ver.col_label"), t("ver.col_created"), t("ver.col_actions")])
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        layout.addWidget(self.table)

        # Snapshot detail
        detail_group = QGroupBox(t("ver.detail_group"))
        detail_layout = QVBoxLayout(detail_group)
        self.detail_text = QTextEdit()
        self.detail_text.setReadOnly(True)
        self.detail_text.setMaximumHeight(200)
        detail_layout.addWidget(self.detail_text)
        layout.addWidget(detail_group)

    def set_project(self, project_id):
        self.project_id = project_id
        self.refresh()

    def refresh(self):
        if not self.project_id:
            return
        versions = self.db.get_versions(self.project_id)
        self.table.setRowCount(len(versions))
        for i, v in enumerate(versions):
            self.table.setItem(i, 0, QTableWidgetItem(str(v['id'])))
            self.table.setItem(i, 1, QTableWidgetItem(v['label'] or t("ver.auto_save_label", id=v['id'])))
            self.table.setItem(i, 2, QTableWidgetItem(v['created_at']))

            # Action buttons
            btn_layout = QWidget()
            bl = QHBoxLayout(btn_layout)
            bl.setContentsMargins(2, 2, 2, 2)

            btn_view = QPushButton(t("ver.view_btn"))
            btn_view.setObjectName("flat")
            btn_view.clicked.connect(lambda _, vid=v['id']: self._view_version(vid))
            bl.addWidget(btn_view)

            btn_restore = QPushButton(t("ver.restore_btn"))
            btn_restore.setObjectName("flat")
            btn_restore.setStyleSheet("color: #E74C3C;")
            btn_restore.clicked.connect(lambda _, vid=v['id']: self._restore_version(vid))
            bl.addWidget(btn_restore)

            self.table.setCellWidget(i, 3, btn_layout)

    def _save_version(self):
        if not self.project_id:
            return
        label = self.label_input.text().strip()
        self.db.save_version(self.project_id, label)
        self.label_input.clear()
        self.refresh()
        QMessageBox.information(self, t("common.success"), t("ver.saved_success"))

    def _view_version(self, version_id):
        snapshot = self.db.get_version_snapshot(version_id)
        if not snapshot:
            return
        info = t("ver.detail_project", v=snapshot.get('project', {}).get('name', ''))
        info += t("ver.detail_voc_count", v=len(snapshot.get('vocs', [])))
        info += t("ver.detail_ctq_count", v=len(snapshot.get('ctqs', [])))
        info += t("ver.detail_rel_count", v=len(snapshot.get('relationships', [])))
        info += t("ver.detail_roof_count", v=len(snapshot.get('roof', [])))
        info += t("ver.detail_comp_count", v=len(snapshot.get('competitors', [])))
        self.detail_text.setPlainText(info)

    def _restore_version(self, version_id):
        reply = QMessageBox.question(self, t("ver.confirm_restore_title"), t("ver.confirm_restore_body"),
            QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        snapshot = self.db.get_version_snapshot(version_id)
        if not snapshot:
            return

        # Clear current data
        pid = self.project_id
        self.db.conn.execute("DELETE FROM voc WHERE project_id=?", (pid,))
        self.db.conn.execute("DELETE FROM ctq WHERE project_id=?", (pid,))
        self.db.conn.execute("DELETE FROM hoq_relationship WHERE project_id=?", (pid,))
        self.db.conn.execute("DELETE FROM hoq_roof WHERE project_id=?", (pid,))
        self.db.conn.execute("DELETE FROM competitor WHERE project_id=?", (pid,))

        # Restore VOCs
        for v in snapshot.get('vocs', []):
            self.db.conn.execute("""
                INSERT INTO voc (id, project_id, phase, parent_id, name, description,
                source, importance, kano_type, sort_order, improvement_ratio, sales_point, planned_level)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (v['id'], pid, v['phase'], v.get('parent_id'),
                  v['name'], v.get('description', ''),
                  v.get('source', ''), v.get('importance', 3),
                  v.get('kano_type', ''), v.get('sort_order', 0),
                  v.get('improvement_ratio', 1.0), v.get('sales_point', 1.0),
                  v.get('planned_level', 0)))

        # Restore CTQs
        for c in snapshot.get('ctqs', []):
            self.db.conn.execute("""
                INSERT INTO ctq (id, project_id, phase, name, unit, target_value,
                current_value, direction, difficulty, sort_order)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (c['id'], pid, c['phase'], c['name'],
                  c.get('unit', ''), c.get('target_value', ''),
                  c.get('current_value', ''), c.get('direction', 'higher_better'),
                  c.get('difficulty', 3), c.get('sort_order', 0)))

        # Restore relationships
        for r in snapshot.get('relationships', []):
            self.db.conn.execute("""
                INSERT INTO hoq_relationship (project_id, phase, voc_id, ctq_id, strength, symbol)
                VALUES (?,?,?,?,?,?)
            """, (pid, r['phase'], r['voc_id'], r['ctq_id'],
                  r['strength'], r.get('symbol', '')))

        # Restore roof
        for rf in snapshot.get('roof', []):
            self.db.conn.execute("""
                INSERT INTO hoq_roof (project_id, phase, ctq_id_1, ctq_id_2, correlation)
                VALUES (?,?,?,?,?)
            """, (pid, rf['phase'], rf['ctq_id_1'], rf['ctq_id_2'], rf['correlation']))

        self.db.conn.commit()
        self.version_restored.emit()
        QMessageBox.information(self, t("common.success"), t("ver.restored_success"))
