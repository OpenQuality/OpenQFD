"""
VOC (Voice of Customer) Management View
Supports tree structure (3-level nesting), CRUD, Kano classification,
importance rating, batch import from CSV/Excel.
"""
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTreeWidget, QTreeWidgetItem,
    QPushButton, QLabel, QLineEdit, QComboBox, QDoubleSpinBox,
    QTextEdit, QGroupBox, QFormLayout, QMessageBox, QFileDialog,
    QHeaderView, QAbstractItemView, QSplitter, QDialog, QDialogButtonBox,
    QSpinBox, QFrame
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QColor, QFont, QIcon
import csv
import os

from views.styles import KANO_TYPES, KANO_TYPE_KEYS, COLORS
from utils.i18n import t


class VOCManagerView(QWidget):
    data_changed = Signal()

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.project_id = None
        self.phase = 1
        self._setup_ui()

    def _setup_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        splitter = QSplitter(Qt.Horizontal)

        # Left: Tree view
        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(0, 0, 0, 0)

        header_layout = QHBoxLayout()
        lbl = QLabel(t("voc.title"))
        lbl.setObjectName("title")
        header_layout.addWidget(lbl)
        header_layout.addStretch()

        btn_add = QPushButton(t("voc.add"))
        btn_add.clicked.connect(self._add_voc)
        header_layout.addWidget(btn_add)

        btn_add_child = QPushButton(t("voc.add_child"))
        btn_add_child.setObjectName("secondary")
        btn_add_child.clicked.connect(self._add_child_voc)
        header_layout.addWidget(btn_add_child)

        btn_import = QPushButton(t("voc.import"))
        btn_import.setObjectName("accent")
        btn_import.clicked.connect(self._import_voc)
        header_layout.addWidget(btn_import)

        btn_template = QPushButton(t("voc.download_template"))
        btn_template.setObjectName("secondary")
        btn_template.clicked.connect(self._download_template)
        header_layout.addWidget(btn_template)

        left_layout.addLayout(header_layout)

        # Filter
        filter_layout = QHBoxLayout()
        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText(t("voc.search"))
        self.filter_input.textChanged.connect(self._filter_tree)
        filter_layout.addWidget(self.filter_input)

        self.kano_filter = QComboBox()
        self.kano_filter.addItem(t("voc.all_kano"), "")
        for k in KANO_TYPES:
            if k:
                self.kano_filter.addItem(t(KANO_TYPE_KEYS[k]), k)
        self.kano_filter.currentIndexChanged.connect(self._filter_tree)
        self.kano_filter.setFixedWidth(200)
        filter_layout.addWidget(self.kano_filter)
        left_layout.addLayout(filter_layout)

        self.tree = QTreeWidget()
        self.tree.setHeaderLabels([t("voc.col_name"), t("voc.col_importance"), t("voc.col_kano"),
                                    t("voc.col_source"), t("voc.col_sales_point"), t("voc.col_ui"), t("voc.col_ti")])
        self.tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
        for i in range(1, 7):
            self.tree.header().setSectionResizeMode(i, QHeaderView.ResizeToContents)
        self.tree.setAlternatingRowColors(True)
        self.tree.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tree.currentItemChanged.connect(self._on_selection_changed)
        self.tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tree.setDragDropMode(QAbstractItemView.InternalMove)
        left_layout.addWidget(self.tree)

        # Bottom buttons
        bottom_layout = QHBoxLayout()
        btn_del = QPushButton(t("common.delete_selected"))
        btn_del.setObjectName("danger")
        btn_del.clicked.connect(self._delete_voc)
        bottom_layout.addWidget(btn_del)

        btn_up = QPushButton(t("common.up"))
        btn_up.setObjectName("secondary")
        btn_up.clicked.connect(lambda: self._move_voc(-1))
        bottom_layout.addWidget(btn_up)

        btn_down = QPushButton(t("common.down"))
        btn_down.setObjectName("secondary")
        btn_down.clicked.connect(lambda: self._move_voc(1))
        bottom_layout.addWidget(btn_down)

        bottom_layout.addStretch()
        self.count_label = QLabel(t("voc.count", n=0))
        self.count_label.setObjectName("subtitle")
        bottom_layout.addWidget(self.count_label)
        left_layout.addLayout(bottom_layout)

        splitter.addWidget(left)

        # Right: Detail editor
        right = QWidget()
        right.setMinimumWidth(320)
        right.setMaximumWidth(420)
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(0, 0, 0, 0)

        detail_label = QLabel(t("voc.detail"))
        detail_label.setObjectName("title")
        right_layout.addWidget(detail_label)

        form_group = QGroupBox(t("voc.basic_info"))
        form = QFormLayout(form_group)

        self.edit_name = QLineEdit()
        self.edit_name.setPlaceholderText(t("voc.name_placeholder"))
        form.addRow(t("voc.name"), self.edit_name)

        self.edit_desc = QTextEdit()
        self.edit_desc.setMaximumHeight(80)
        self.edit_desc.setPlaceholderText(t("voc.desc_placeholder"))
        form.addRow(t("voc.desc"), self.edit_desc)

        self.edit_source = QComboBox()
        self.edit_source.setEditable(True)
        self.edit_source.addItems(["", t("src.research"), t("src.interview"), t("src.complaint"),
                                    t("src.standard"), t("src.benchmark"), t("src.internal")])
        form.addRow(t("voc.source"), self.edit_source)

        right_layout.addWidget(form_group)

        weight_group = QGroupBox(t("voc.weight_settings"))
        weight_form = QFormLayout(weight_group)

        self.edit_importance = QDoubleSpinBox()
        self.edit_importance.setRange(0.0, 10.0)
        self.edit_importance.setSingleStep(0.1)
        self.edit_importance.setDecimals(2)
        self.edit_importance.setValue(3.0)
        self.edit_importance.setToolTip(t("voc.importance_tooltip"))
        self.edit_importance.valueChanged.connect(self._update_ri_display)
        weight_form.addRow(t("voc.importance"), self.edit_importance)

        self.edit_kano = QComboBox()
        for k in KANO_TYPES:
            self.edit_kano.addItem(t(KANO_TYPE_KEYS[k]), k)
        weight_form.addRow(t("voc.kano"), self.edit_kano)

        self.edit_sales_point = QDoubleSpinBox()
        self.edit_sales_point.setRange(0.1, 3.0)
        self.edit_sales_point.setSingleStep(0.1)
        self.edit_sales_point.setDecimals(1)
        self.edit_sales_point.setValue(1.0)
        self.edit_sales_point.setToolTip(t("voc.sales_point_tooltip"))
        self.edit_sales_point.valueChanged.connect(self._update_ri_display)
        weight_form.addRow(t("voc.sales_point"), self.edit_sales_point)

        self.edit_ui = QDoubleSpinBox()
        self.edit_ui.setRange(0.0, 10.0)
        self.edit_ui.setSingleStep(0.5)
        self.edit_ui.setDecimals(1)
        self.edit_ui.setToolTip(t("voc.ui_tooltip"))
        self.edit_ui.valueChanged.connect(self._update_ri_display)
        weight_form.addRow(t("voc.ui_label"), self.edit_ui)

        self.edit_planned = QDoubleSpinBox()
        self.edit_planned.setRange(0.0, 10.0)
        self.edit_planned.setSingleStep(0.5)
        self.edit_planned.setDecimals(1)
        self.edit_planned.valueChanged.connect(self._update_ri_display)
        weight_form.addRow(t("voc.ti_label"), self.edit_planned)

        self.ri_label = QLabel("-")
        self.ri_label.setStyleSheet("color: #1B5E8C; font-weight: bold;")
        self.ri_label.setToolTip(t("voc.ri_tooltip"))
        weight_form.addRow(t("voc.ri_label_text"), self.ri_label)

        right_layout.addWidget(weight_group)

        # Adjusted weight display
        self.adj_weight_label = QLabel(t("voc.adj_weight_full", v="-"))
        self.adj_weight_label.setStyleSheet("color: #1B5E8C; font-weight: bold; font-size: 14px; padding: 8px;")
        right_layout.addWidget(self.adj_weight_label)

        save_btn = QPushButton(t("voc.save"))
        save_btn.setObjectName("success")
        save_btn.setMinimumHeight(40)
        save_btn.clicked.connect(self._save_current)
        right_layout.addWidget(save_btn)

        right_layout.addStretch()
        splitter.addWidget(right)

        splitter.setSizes([600, 350])
        layout.addWidget(splitter)

    def set_project(self, project_id, phase=1):
        self.project_id = project_id
        self.phase = phase
        self.refresh()

    def refresh(self):
        if not self.project_id:
            return
        self.tree.clear()
        vocs = self.db.get_vocs(self.project_id, self.phase)

        # Build tree from flat list
        item_map = {}
        roots = []
        for v in vocs:
            item = QTreeWidgetItem()
            item.setData(0, Qt.UserRole, v['id'])
            item.setText(0, v['name'])
            item.setText(1, f"{v['importance']:.2f}")
            item.setText(2, t(KANO_TYPE_KEYS.get(v['kano_type'], 'kano.unclassified')))
            item.setText(3, v['source'] or '')
            item.setText(4, f"{v['sales_point']:.1f}")
            item.setText(5, f"{v['current_level'] or 0:.1f}")
            item.setText(6, f"{v['planned_level'] or 0:.1f}")

            # Color code Kano type
            kano_colors = {
                'M': '#FADBD8', 'O': '#D4E6F1', 'A': '#D5F5E3',
                'I': '#E5E8E8', 'R': '#FDEBD0'
            }
            if v['kano_type'] in kano_colors:
                item.setBackground(2, QColor(kano_colors[v['kano_type']]))

            item_map[v['id']] = item
            if v['parent_id'] and v['parent_id'] in item_map:
                item_map[v['parent_id']].addChild(item)
            else:
                roots.append(item)

        for r in roots:
            self.tree.addTopLevelItem(r)
        self.tree.expandAll()

        self.count_label.setText(t("voc.count", n=len(vocs)))

    def _on_selection_changed(self, current, previous):
        if not current:
            return
        voc_id = current.data(0, Qt.UserRole)
        if not voc_id:
            return
        rows = self.db.conn.execute("SELECT * FROM voc WHERE id=?", (voc_id,)).fetchone()
        if not rows:
            return

        self.edit_name.setText(rows['name'])
        self.edit_desc.setPlainText(rows['description'] or '')
        idx = self.edit_source.findText(rows['source'] or '')
        if idx >= 0:
            self.edit_source.setCurrentIndex(idx)
        else:
            self.edit_source.setEditText(rows['source'] or '')
        self.edit_importance.setValue(rows['importance'] or 3.0)
        kano_idx = self.edit_kano.findData(rows['kano_type'] or '')
        if kano_idx >= 0:
            self.edit_kano.setCurrentIndex(kano_idx)
        self.edit_sales_point.setValue(rows['sales_point'] or 1.0)
        self.edit_ui.setValue(rows['current_level'] or 0.0)
        self.edit_planned.setValue(rows['planned_level'] or 0.0)
        self._update_ri_display()

    def _update_ri_display(self):
        """Ri = Ti/Ui is derived, not entered — recompute it (and the Wai preview
        that depends on it) whenever Ui/Ti/Ii/Si change."""
        ui = self.edit_ui.value()
        ti = self.edit_planned.value()
        ri = (ti / ui) if ui > 0 else 0
        self.ri_label.setText(f"{ri:.2f}" if ui > 0 else t("voc.ri_needs_ui"))

        si = self.edit_sales_point.value()
        ii = self.edit_importance.value()
        wai = ri * si * ii
        self.adj_weight_label.setText(t("voc.adj_weight_full", v=f"{wai:.2f}"))

    def _save_current(self):
        item = self.tree.currentItem()
        if not item:
            QMessageBox.warning(self, t("common.hint"), t("voc.select_first"))
            return
        voc_id = item.data(0, Qt.UserRole)
        self.db.update_voc(voc_id,
            name=self.edit_name.text(),
            description=self.edit_desc.toPlainText(),
            source=self.edit_source.currentText(),
            importance=self.edit_importance.value(),
            kano_type=self.edit_kano.currentData(),
            sales_point=self.edit_sales_point.value(),
            planned_level=self.edit_planned.value(),
        )
        # Ui goes through the sync helper so CBA's self-competitor score and
        # HOQ's Ui column stay identical to what's entered here.
        self.db.sync_self_voc_score(self.project_id, voc_id, self.edit_ui.value())
        self.refresh()
        self.data_changed.emit()

    def _add_voc(self):
        if not self.project_id:
            return
        voc_id = self.db.add_voc(self.project_id, t("voc.new_item_name"), phase=self.phase)
        self.refresh()
        self.data_changed.emit()
        # Select the new item
        self._select_by_id(voc_id)

    def _add_child_voc(self):
        item = self.tree.currentItem()
        if not item:
            QMessageBox.information(self, t("common.hint"), t("voc.select_parent_first"))
            return
        parent_id = item.data(0, Qt.UserRole)
        # Check nesting level (max 3)
        level = 0
        p = item
        while p.parent():
            level += 1
            p = p.parent()
        if level >= 2:
            QMessageBox.warning(self, t("voc.limit_title"), t("voc.max_nesting"))
            return
        voc_id = self.db.add_voc(self.project_id, t("voc.new_child_name"), phase=self.phase, parent_id=parent_id)
        self.refresh()
        self.data_changed.emit()
        self._select_by_id(voc_id)

    def _delete_voc(self):
        item = self.tree.currentItem()
        if not item:
            return
        voc_id = item.data(0, Qt.UserRole)
        reply = QMessageBox.question(self, t("msg.confirm_delete_title"),
            t("voc.confirm_delete_body", name=item.text(0)),
            QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            # Delete children first
            self._delete_children(voc_id)
            self.db.delete_voc(voc_id)
            self.refresh()
            self.data_changed.emit()

    def _delete_children(self, parent_id):
        children = self.db.conn.execute(
            "SELECT id FROM voc WHERE parent_id=?", (parent_id,)).fetchall()
        for c in children:
            self._delete_children(c['id'])
            self.db.delete_voc(c['id'])

    def _move_voc(self, direction):
        item = self.tree.currentItem()
        if not item:
            return
        voc_id = item.data(0, Qt.UserRole)
        vocs = self.db.get_vocs(self.project_id, self.phase)
        ids = [v['id'] for v in vocs]
        if voc_id not in ids:
            return
        idx = ids.index(voc_id)
        new_idx = idx + direction
        if 0 <= new_idx < len(ids):
            for i, vid in enumerate(ids):
                self.db.update_voc(vid, sort_order=i)
            self.db.update_voc(ids[idx], sort_order=new_idx)
            self.db.update_voc(ids[new_idx], sort_order=idx)
            self.refresh()

    def _import_voc(self):
        path, _ = QFileDialog.getOpenFileName(self, t("voc.import_title"), "", t("voc.import_filter"))
        if not path:
            return
        try:
            if path.endswith('.csv'):
                self._import_csv(path)
            elif path.endswith('.xlsx'):
                self._import_xlsx(path)
            self.refresh()
            self.data_changed.emit()
            QMessageBox.information(self, t("common.success"), t("common.import_done"))
        except Exception as e:
            QMessageBox.critical(self, t("msg.import_failed"), str(e))

    def _row_to_voc_kwargs(self, row):
        """Extract recognized VOC fields from an imported row (a dict-like mapping
        of column header -> value). Returns (name, kwargs, ui_or_None)."""
        def pick(*keys, cast=None):
            for k in keys:
                val = row.get(k)
                if val not in (None, ''):
                    return cast(val) if cast else val
            return None

        name = pick('name', '需求名称', '需求')
        kwargs = {}
        v = pick('重要度', 'importance', cast=float)
        if v is not None: kwargs['importance'] = v
        v = pick('来源', 'source')
        if v is not None: kwargs['source'] = str(v)
        v = pick('描述', 'description')
        if v is not None: kwargs['description'] = str(v)
        v = pick('kano', 'Kano', 'kano_type')
        if v is not None: kwargs['kano_type'] = str(v)
        v = pick('销售点', 'sales_point', cast=float)
        if v is not None: kwargs['sales_point'] = v
        v = pick('计划水平', 'Ti', 'planned_level', cast=float)
        if v is not None: kwargs['planned_level'] = v
        ui = pick('当前水平', 'Ui', 'current_level', cast=float)
        return name, kwargs, ui

    def _import_csv(self, path):
        with open(path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                name, kwargs, ui = self._row_to_voc_kwargs(row)
                if not name:
                    continue
                voc_id = self.db.add_voc(self.project_id, name, phase=self.phase, **kwargs)
                if ui is not None:
                    self.db.sync_self_voc_score(self.project_id, voc_id, ui)

    def _import_xlsx(self, path):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {k: ('' if v is None else v) for k, v in zip(headers, row)}
            name, kwargs, ui = self._row_to_voc_kwargs(data)
            if not name:
                continue
            voc_id = self.db.add_voc(self.project_id, str(name), phase=self.phase, **kwargs)
            if ui is not None:
                self.db.sync_self_voc_score(self.project_id, voc_id, ui)
        wb.close()

    def _download_template(self):
        path, _ = QFileDialog.getSaveFileName(
            self, t("voc.download_template_title"), t("voc.template_filename"), t("common.csv_filter"))
        if not path:
            return
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow([t("voc.tpl_name"), t("voc.tpl_desc"), t("voc.tpl_source"), t("voc.tpl_importance"),
                              t("voc.tpl_kano"), t("voc.tpl_sales_point"), t("voc.tpl_ui"), t("voc.tpl_ti")])
            writer.writerow(["电池续航长", "希望一次充电可用两天以上", "客户访谈", "5", "M", "1.5", "3", "4.5"])
            writer.writerow(["屏幕显示清晰", "", "市场调研", "4", "O", "1.0", "3.5", "4"])
        QMessageBox.information(self, t("common.success"), t("voc.template_saved_body", path=path))

    def _filter_tree(self):
        text = self.filter_input.text().lower()
        kano = self.kano_filter.currentData()
        for i in range(self.tree.topLevelItemCount()):
            self._filter_item(self.tree.topLevelItem(i), text, kano)

    def _filter_item(self, item, text, kano):
        match = True
        if text and text not in item.text(0).lower():
            match = False
        if kano and kano not in item.text(2):
            match = False
        # Check children
        child_visible = False
        for i in range(item.childCount()):
            if self._filter_item(item.child(i), text, kano):
                child_visible = True
        visible = match or child_visible
        item.setHidden(not visible)
        return visible

    def _select_by_id(self, voc_id):
        iterator = self.tree.findItems("", Qt.MatchContains | Qt.MatchRecursive, 0)
        for item in iterator:
            if item.data(0, Qt.UserRole) == voc_id:
                self.tree.setCurrentItem(item)
                break
